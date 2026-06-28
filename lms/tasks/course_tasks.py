"""
Celery Tasks - Course Related
4 Tasks utama:
1. send_enrollment_email     - Email saat student enroll
2. generate_certificate      - Generate certificate saat course complete
3. update_course_statistics  - Update enrollment count (scheduled)
4. export_course_report      - Generate CSV report (async)
"""
import csv
import io
import logging
from datetime import datetime
from celery import shared_task
from django.conf import settings
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.contrib.auth.models import User

logger = logging.getLogger(__name__)


# ============================================================
# TASK 1: SEND ENROLLMENT EMAIL
# ============================================================
@shared_task(
    bind=True,
    name='lms.tasks.course_tasks.send_enrollment_email',
    max_retries=3,
    default_retry_delay=60,  # Retry setelah 60 detik
    autoretry_for=(Exception,),
)
def send_enrollment_email(self, enrollment_id):
    """
    Kirim email konfirmasi enrollment ke student.

    Args:
        enrollment_id: ID enrollment di PostgreSQL

    Flow:
    1. Ambil data enrollment dari database
    2. Generate email content
    3. Kirim email via SMTP
    4. Log ke MongoDB
    5. Invalidate cache jika perlu
    """
    try:
        logger.info(f"[TASK] send_enrollment_email: enrollment_id={enrollment_id}")

        from lms.apps.courses.models import Enrollment
        enrollment = Enrollment.objects.select_related(
            'student', 'course', 'course__instructor'
        ).get(id=enrollment_id)

        student = enrollment.student
        course = enrollment.course

        # Buat email content
        subject = f"Selamat! Anda Terdaftar di Kursus: {course.title}"

        message = f"""
Halo {student.get_full_name() or student.username},

Selamat! Anda berhasil mendaftar di kursus berikut:

📚 Nama Kursus : {course.title}
👨‍🏫 Instruktur  : {course.instructor.get_full_name()}
📅 Tanggal Daftar: {enrollment.enrolled_at.strftime('%d %B %Y')}

Selamat belajar dan semoga sukses!

Salam,
Tim LMS
        """

        # Kirim email
        send_mail(
            subject=subject,
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[student.email],
            fail_silently=False,
        )

        logger.info(f"[TASK] Email enrollment terkirim ke {student.email}")

        # Log aktivitas ke MongoDB
        _log_task_activity(
            user_id=student.id,
            username=student.username,
            action='course_enroll',
            resource_type='course',
            resource_id=course.id,
            resource_name=course.title,
            metadata={'enrollment_id': enrollment_id, 'email_sent': True}
        )

        return {
            'status': 'success',
            'enrollment_id': enrollment_id,
            'email': student.email,
        }

    except Enrollment.DoesNotExist:
        logger.error(f"[TASK] Enrollment {enrollment_id} tidak ditemukan")
        return {'status': 'error', 'message': 'Enrollment not found'}

    except Exception as exc:
        logger.error(f"[TASK] Error send_enrollment_email: {exc}")
        raise self.retry(exc=exc)


# ============================================================
# TASK 2: GENERATE CERTIFICATE
# ============================================================
@shared_task(
    bind=True,
    name='lms.tasks.course_tasks.generate_certificate',
    max_retries=3,
    default_retry_delay=120,
)
def generate_certificate(self, enrollment_id):
    """
    Generate sertifikat penyelesaian kursus menggunakan ReportLab.

    Args:
        enrollment_id: ID enrollment yang telah selesai (progress = 100%)

    Flow:
    1. Validasi enrollment sudah complete
    2. Generate PDF certificate dengan ReportLab
    3. Simpan file ke media storage
    4. Update Enrollment.certificate_issued = True
    5. Kirim email dengan attachment certificate
    6. Invalidate cache
    """
    try:
        logger.info(f"[TASK] generate_certificate: enrollment_id={enrollment_id}")

        from lms.apps.courses.models import Enrollment
        enrollment = Enrollment.objects.select_related(
            'student', 'course', 'course__instructor'
        ).get(id=enrollment_id)

        # Validasi completion
        if enrollment.progress < 100.0:
            logger.warning(
                f"[TASK] Enrollment {enrollment_id} belum complete "
                f"(progress: {enrollment.progress}%)"
            )
            return {'status': 'skipped', 'reason': 'Course not completed yet'}

        # Generate PDF certificate
        pdf_content = _generate_certificate_pdf(enrollment)

        # Simpan ke file system
        import os
        cert_dir = os.path.join(settings.MEDIA_ROOT, 'certificates')
        os.makedirs(cert_dir, exist_ok=True)

        cert_filename = f"certificate_{enrollment.student.username}_{enrollment.course.slug}.pdf"
        cert_path = os.path.join(cert_dir, cert_filename)

        with open(cert_path, 'wb') as f:
            f.write(pdf_content)

        # Update enrollment record
        from django.utils import timezone
        enrollment.certificate_issued = True
        enrollment.completed_at = timezone.now()
        enrollment.status = 'completed'
        enrollment.save(update_fields=['certificate_issued', 'completed_at', 'status'])

        # Kirim email dengan certificate
        _send_certificate_email(enrollment, cert_path)

        # Invalidate course cache
        from lms.apps.analytics.cache_service import CacheService
        CacheService.invalidate_course_cache(enrollment.course.id)

        # Log ke MongoDB
        _log_task_activity(
            user_id=enrollment.student.id,
            username=enrollment.student.username,
            action='course_complete',
            resource_type='course',
            resource_id=enrollment.course.id,
            resource_name=enrollment.course.title,
            metadata={
                'enrollment_id': enrollment_id,
                'certificate_path': cert_filename,
                'completed_at': timezone.now().isoformat(),
            }
        )

        logger.info(f"[TASK] Certificate generated: {cert_filename}")
        return {
            'status': 'success',
            'enrollment_id': enrollment_id,
            'certificate_file': cert_filename,
        }

    except Exception as exc:
        logger.error(f"[TASK] Error generate_certificate: {exc}")
        raise self.retry(exc=exc)


def _generate_certificate_pdf(enrollment):
    """Generate PDF certificate menggunakan ReportLab."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=72, leftMargin=72,
        topMargin=72, bottomMargin=72
    )

    styles = getSampleStyleSheet()
    story = []

    # Title style
    title_style = ParagraphStyle(
        'CertTitle',
        parent=styles['Title'],
        fontSize=36,
        spaceAfter=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#1a3a5c'),
    )

    subtitle_style = ParagraphStyle(
        'CertSubtitle',
        parent=styles['Normal'],
        fontSize=18,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#555555'),
    )

    name_style = ParagraphStyle(
        'CertName',
        parent=styles['Normal'],
        fontSize=28,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#2e75b6'),
        spaceAfter=20,
        fontName='Helvetica-Bold',
    )

    body_style = ParagraphStyle(
        'CertBody',
        parent=styles['Normal'],
        fontSize=14,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#333333'),
    )

    # Build content
    story.append(Spacer(1, 0.5 * inch))
    story.append(Paragraph("SERTIFIKAT PENYELESAIAN", title_style))
    story.append(Paragraph("Certificate of Completion", subtitle_style))
    story.append(Spacer(1, 0.5 * inch))
    story.append(Paragraph("Diberikan kepada:", body_style))
    story.append(Spacer(1, 0.2 * inch))
    story.append(Paragraph(
        enrollment.student.get_full_name() or enrollment.student.username,
        name_style
    ))
    story.append(Paragraph("Atas keberhasilan menyelesaikan kursus:", body_style))
    story.append(Spacer(1, 0.2 * inch))
    story.append(Paragraph(enrollment.course.title, name_style))
    story.append(Spacer(1, 0.3 * inch))

    completion_date = enrollment.completed_at or datetime.now()
    story.append(Paragraph(
        f"Diselesaikan pada: {completion_date.strftime('%d %B %Y')}",
        body_style
    ))
    story.append(Spacer(1, 0.2 * inch))
    story.append(Paragraph(
        f"Instruktur: {enrollment.course.instructor.get_full_name()}",
        body_style
    ))

    doc.build(story)
    return buffer.getvalue()


def _send_certificate_email(enrollment, cert_path):
    """Kirim email dengan certificate attachment."""
    from django.core.mail import EmailMessage

    subject = f"Sertifikat Kursus: {enrollment.course.title}"
    message = (
        f"Selamat {enrollment.student.get_full_name() or enrollment.student.username}!\n\n"
        f"Anda telah berhasil menyelesaikan kursus '{enrollment.course.title}'.\n"
        f"Sertifikat terlampir dalam email ini.\n\n"
        f"Terima kasih telah belajar bersama kami!\n\nTim LMS"
    )

    email = EmailMessage(
        subject=subject,
        body=message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[enrollment.student.email],
    )
    email.attach_file(cert_path)
    email.send(fail_silently=True)


# ============================================================
# TASK 3: UPDATE COURSE STATISTICS (SCHEDULED)
# ============================================================
@shared_task(
    name='lms.tasks.course_tasks.update_course_statistics',
    bind=True,
)
def update_course_statistics(self):
    """
    Update enrollment count dan statistik kursus secara berkala.
    Dijadwalkan setiap 30 menit oleh Celery Beat.

    Flow:
    1. Ambil semua course yang published
    2. Hitung enrollment count dari database
    3. Update field enrollment_count di PostgreSQL
    4. Update CourseAnalytics di MongoDB
    5. Invalidate Redis cache untuk course yang berubah
    """
    try:
        logger.info("[TASK] update_course_statistics: Starting scheduled task...")

        from lms.apps.courses.models import Course, Enrollment
        from lms.apps.analytics.documents import CourseAnalytics
        from lms.apps.analytics.cache_service import CacheService
        from django.db.models import Count, Avg

        updated_count = 0
        courses = Course.objects.filter(status='published')

        for course in courses:
            # Hitung statistik dari PostgreSQL
            stats = Enrollment.objects.filter(course=course).aggregate(
                total=Count('id'),
                active=Count('id', filter=__import__('django.db.models', fromlist=['Q']).Q(status='active')),
                completed=Count('id', filter=__import__('django.db.models', fromlist=['Q']).Q(status='completed')),
                avg_progress=Avg('progress'),
            )

            total = stats['total'] or 0
            completed = stats['completed'] or 0
            avg_progress = stats['avg_progress'] or 0.0

            # Update PostgreSQL enrollment_count
            old_count = course.enrollment_count
            course.enrollment_count = total
            course.save(update_fields=['enrollment_count'])

            # Update MongoDB analytics
            try:
                completion_rate = (completed / total * 100) if total > 0 else 0.0
                CourseAnalytics.objects(course_id=course.id).update_one(
                    set__course_id=course.id,
                    set__course_title=course.title,
                    set__total_enrollments=total,
                    set__active_students=stats['active'] or 0,
                    set__completed_students=completed,
                    set__avg_progress=avg_progress,
                    set__completion_rate=completion_rate,
                    set__last_updated=datetime.utcnow(),
                    upsert=True
                )
            except Exception as mongo_err:
                logger.error(f"MongoDB update error for course {course.id}: {mongo_err}")

            # Invalidate cache jika ada perubahan
            if old_count != total:
                CacheService.invalidate_course_cache(course.id)
                updated_count += 1

        logger.info(f"[TASK] update_course_statistics: {updated_count} courses updated")
        return {
            'status': 'success',
            'courses_processed': courses.count(),
            'courses_updated': updated_count,
            'timestamp': datetime.now().isoformat(),
        }

    except Exception as exc:
        logger.error(f"[TASK] Error update_course_statistics: {exc}")
        raise


# ============================================================
# TASK 4: EXPORT COURSE REPORT (ASYNC)
# ============================================================
@shared_task(
    bind=True,
    name='lms.tasks.course_tasks.export_course_report',
    max_retries=2,
    soft_time_limit=300,  # 5 menit timeout
    time_limit=360,
)
def export_course_report(self, course_id, requested_by_user_id, report_format='csv'):
    """
    Generate laporan enrollment course secara asynchronous.
    Mendukung format CSV dan Excel.

    Args:
        course_id: ID course
        requested_by_user_id: User yang request report
        report_format: 'csv' atau 'excel'

    Flow:
    1. Ambil data enrollment + student dari PostgreSQL
    2. Generate file report (CSV/Excel)
    3. Simpan ke media storage
    4. Kirim notifikasi email ke requester
    5. Return path file untuk download
    """
    try:
        logger.info(
            f"[TASK] export_course_report: course_id={course_id}, "
            f"format={report_format}, requested_by={requested_by_user_id}"
        )

        from lms.apps.courses.models import Course, Enrollment

        course = Course.objects.get(id=course_id)
        requester = User.objects.get(id=requested_by_user_id)

        # Ambil data enrollment
        enrollments = Enrollment.objects.filter(
            course=course
        ).select_related('student').order_by('-enrolled_at')

        # Build data rows
        rows = []
        for enrollment in enrollments:
            student = enrollment.student
            rows.append({
                'Student ID': student.id,
                'Username': student.username,
                'Nama Lengkap': student.get_full_name(),
                'Email': student.email,
                'Status Enrollment': enrollment.get_status_display(),
                'Progress (%)': f"{enrollment.progress:.1f}",
                'Tanggal Daftar': enrollment.enrolled_at.strftime('%d/%m/%Y %H:%M'),
                'Tanggal Selesai': (
                    enrollment.completed_at.strftime('%d/%m/%Y %H:%M')
                    if enrollment.completed_at else '-'
                ),
                'Sertifikat': 'Ya' if enrollment.certificate_issued else 'Tidak',
            })

        # Generate file berdasarkan format
        import os
        report_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(report_dir, exist_ok=True)

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename_base = f"report_{course.slug}_{timestamp}"

        if report_format == 'excel':
            filepath = _generate_excel_report(rows, course, report_dir, filename_base)
        else:
            filepath = _generate_csv_report(rows, report_dir, filename_base)

        # Kirim email notifikasi
        relative_path = os.path.relpath(filepath, settings.MEDIA_ROOT)
        download_url = f"{settings.MEDIA_URL}{relative_path}"

        send_mail(
            subject=f"Laporan Kursus Siap: {course.title}",
            message=(
                f"Halo {requester.get_full_name() or requester.username},\n\n"
                f"Laporan untuk kursus '{course.title}' telah selesai diproses.\n"
                f"Jumlah data: {len(rows)} enrollment\n\n"
                f"File tersedia di: {download_url}\n\n"
                f"Salam,\nTim LMS"
            ),
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[requester.email],
            fail_silently=True,
        )

        logger.info(f"[TASK] Report generated: {filepath}")
        return {
            'status': 'success',
            'course_id': course_id,
            'filepath': filepath,
            'download_url': download_url,
            'total_rows': len(rows),
        }

    except Exception as exc:
        logger.error(f"[TASK] Error export_course_report: {exc}")
        raise self.retry(exc=exc)


def _generate_csv_report(rows, report_dir, filename_base):
    """Generate CSV report file."""
    import os
    filepath = os.path.join(report_dir, f"{filename_base}.csv")

    if not rows:
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            f.write("Tidak ada data enrollment")
        return filepath

    with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.DictWriter(f, fieldnames=rows[0].keys())
        writer.writeheader()
        writer.writerows(rows)

    return filepath


def _generate_excel_report(rows, course, report_dir, filename_base):
    """Generate Excel report file menggunakan openpyxl."""
    import os
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    filepath = os.path.join(report_dir, f"{filename_base}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enrollment Report"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1a3a5c")
    header_alignment = Alignment(horizontal='center')

    if rows:
        headers = list(rows[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data.values(), 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-fit columns
        for col_idx in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = 20

    wb.save(filepath)
    return filepath


# ============================================================
# HELPER FUNCTION
# ============================================================
def _log_task_activity(user_id, username, action, resource_type=None,
                        resource_id=None, resource_name=None, metadata=None):
    """Helper untuk log task activity ke MongoDB."""
    try:
        from lms.apps.analytics.documents import ActivityLog
        log = ActivityLog(
            user_id=user_id,
            username=username,
            action=action,
            resource_type=resource_type,
            resource_id=resource_id,
            resource_name=resource_name,
            metadata=metadata or {},
        )
        log.save()
    except Exception as e:
        logger.error(f"Error logging task activity to MongoDB: {e}")
